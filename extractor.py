"""
extractor.py — DataFrame normaliser + styled Excel/JSON export
"""

import io
import json
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def records_to_dataframe(records: list[dict], source_url: str = "") -> pd.DataFrame:
    """Flatten list of dicts into a tidy DataFrame."""
    flat = [_flatten(record) for record in records]
    df = pd.DataFrame(flat)
    if source_url:
        df.insert(0, "source_url", source_url)
    return df


def merge_dataframes(dfs: list[pd.DataFrame]) -> pd.DataFrame:
    """Merge multiple DataFrames into one unified sheet."""
    if not dfs:
        return pd.DataFrame()
    merged = pd.concat(dfs, ignore_index=True, sort=False)
    return merged.fillna("")


def _flatten(d: dict, parent_key: str = "", sep: str = ".") -> dict:
    """Recursively flatten nested dicts."""
    items = {}
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.update(_flatten(v, new_key, sep=sep))
        elif isinstance(v, list):
            items[new_key] = json.dumps(v, ensure_ascii=False)
        else:
            items[new_key] = v
    return items


def dataframe_to_excel_bytes(df: pd.DataFrame, per_site_dfs: list[tuple] = None) -> bytes:
    """Build a styled Excel workbook with All Results + per-site sheets."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Sheet 1: All Results
        df.to_excel(writer, index=False, sheet_name="All Results")
        _style_sheet(writer.sheets["All Results"], header_color="1A3A2A")

        # Per-site sheets
        if per_site_dfs:
            for label, site_df in per_site_dfs:
                sheet_name = _safe_sheet_name(label)
                site_df.to_excel(writer, index=False, sheet_name=sheet_name)
                _style_sheet(writer.sheets[sheet_name], header_color="0D2B1A")

    return buffer.getvalue()


def _style_sheet(ws, header_color: str = "1A3A2A"):
    """Apply hacker-green dark styling to worksheet."""
    header_fill = PatternFill("solid", fgColor=header_color)
    header_font = Font(bold=True, color="00FF41", name="Consolas")
    cell_font   = Font(color="C8FFC8", name="Consolas", size=10)
    row_fill    = PatternFill("solid", fgColor="080C0F")
    alt_fill    = PatternFill("solid", fgColor="0A130D")
    thin_border = Border(
        left=Side(style="thin", color="004400"),
        right=Side(style="thin", color="004400"),
        top=Side(style="thin", color="004400"),
        bottom=Side(style="thin", color="004400"),
    )

    for i, row in enumerate(ws.iter_rows()):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            if i == 0:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.fill = alt_fill if i % 2 == 0 else row_fill
                cell.font = cell_font

    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _safe_sheet_name(label: str) -> str:
    for ch in r"\/*?:[]":
        label = label.replace(ch, "_")
    return label[:31]


def dataframe_to_json_bytes(df: pd.DataFrame) -> bytes:
    records = df.to_dict(orient="records")
    return json.dumps(records, ensure_ascii=False, indent=2).encode("utf-8")
